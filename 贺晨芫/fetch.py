#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
巨量算数（抖音）关键词搜索指数采集 —— 直连 API + AES 解密（无浏览器）
================================================================

直接 POST 到巨量算数后端接口，对响应体中的 AES 密文（data 字段）解密，
提取关键词全部历史日期的搜索指数，写入 Excel（按 关键词+日期 增量去重）。

相比浏览器拦截方案：本脚本纯 requests 直连，不依赖浏览器，
速度更快、更适合定时任务 / 云端运行。

依赖：
    pip install requests pycryptodome openpyxl

合规说明（务必阅读）：
  - 本脚本仅供学习与技术研究使用，请遵守网站服务条款（ToS）。
  - 采集使用【你自己的】登录态 Cookie（msToken 等），不绕过登录、不用于商业滥用。
  - 请合理控制频率（默认批间延迟 + 单批上限），避免对服务造成压力。
  - 签名算法与 AES 密钥均为前端公开值，可能随站点版本变动；如失效，请按脚本注释刷新。

用法：
    python fetch.py                  # 采集内联 KEYWORDS（或同级 keywords.txt）
    python fetch.py --verify         # 试跑单批并打印解密结果，便于上线前确认
    python fetch.py --keywords 小米17Ultra 华为mate70
    python fetch.py --start 20240501 --end 20240510
    python fetch.py --days 365       # 相对今天回推 N 天（默认 HISTORY_DAYS）
"""
import argparse
import base64
import json
import os
import sys
import time
import hashlib
from datetime import datetime, timedelta

try:
    from Crypto.Cipher import AES
    from Crypto.Util.Padding import unpad
except ImportError:
    sys.exit("缺少依赖，请先执行：pip install requests pycryptodome openpyxl")

try:
    import requests
except ImportError:
    sys.exit("缺少依赖，请先执行：pip install requests pycryptodome openpyxl")

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.exit("缺少依赖，请先执行：pip install requests pycryptodome openpyxl")


# ========================== 配置区（请按需修改） ==========================

# 要采集的关键词（内联；同级 keywords.txt 存在时优先使用；--keywords 可覆盖）
KEYWORDS = [
    "小米17Ultra",
]

# 时间窗口（默认相对今天回推，单位天）。可用 --start/--end 或 --days 覆盖。
HISTORY_DAYS = 180

# ★ 登录态 Cookie ★（唯一更新点）
# 浏览器登录 https://trendinsight.oceanengine.com 后，
# 打开 DevTools → Network → 任意 XHR 请求 → 复制 Request Headers 里的完整 Cookie 值，粘贴到下面。
# 必须包含有效 msToken（否则会被风控 / 返回挑战页）。
COOKIE = "PASTE_YOUR_COOKIE_HERE"

# AES 解密配置（默认值来自公开逆向资料；若解密失败，请从当前站点前端 JS 重新提取并替换）
#   —— 怎么找 key/iv：DevTools 搜 "decrypt" / "AES"，在断点处取 key、iv（常为 Base64 或明文串）。
AES_MODE = "CBC"                       # "CBC" 或 "CFB"（两种形态都出现过，二选一）
AES_KEY_CBC = "SjXbYTJb7zXoUToSicUL3A=="   # 形态一：Base64 编码的 key（16/24/32 字节）
AES_IV_CBC = "OekMLjghRg8vlX/PemLc+Q=="    # 形态一：Base64 编码的 iv
AES_KEY_CFB = "anN2bXA2NjYsamlh"           # 形态二：原始字符串 key（16 字节）
AES_IV_CFB = "amlheW91LHFpYW53"            # 形态二：原始字符串 iv

ENDPOINT = "https://trendinsight.oceanengine.com/api/open/index/get_multi_keyword_hot_trend"
APP_NAME = "aweme"        # 抖音=aweme；头条=toutiao；抖音火山版=hotsoon；西瓜=ixigua
REGION = []               # 留空=全国；可填省份/城市编码列表

BATCH_SIZE = 5      # 单批关键词数量（接口单批上限，按需调小）
DELAY = 1.5         # 批次之间的延迟（秒），控制频率
TIMEOUT = 20        # 单次请求超时（秒）

# 是否保留搜索指数为 0 的记录
KEEP_ZERO_RECORDS = False

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_ROOT = os.path.dirname(SCRIPT_DIR)  # skill 根目录（与 scripts/ 同级）

# ========================== 固定输出 / 关键词位置 ==========================
# 为避免「跟随运行目录飘移」，Excel 与关键词都固定在下面这个【项目目录】。
# 如果你把项目挪到别处，只改 PROJECT_DIR 这一行即可（也支持环境变量覆盖）。
PROJECT_DIR = os.environ.get("DOUYIN_PROJECT_DIR", r"C:\Users\ASUS\WorkBuddy\抖音")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "output")        # Excel 固定输出到这里
KEYWORDS_DIR = PROJECT_DIR                              # 关键词文件路径（PROJECT_DIR/keywords.txt）

# 增量状态文件：记录每个关键词已采集到的最新日期，二次运行只补新数据（大幅减少解密量）
STATE_FILE = os.path.join(OUTPUT_DIR, "state.json")
CONCURRENCY_DEFAULT = 1     # 默认串行（最稳）；可用 --concurrency N 开启并发加速首跑
MAX_RETRIES_DEFAULT = 3     # 网络 / 5xx 等瞬态错误的重试次数（风控页面不重试，需换 Cookie）
PER_REQUEST_DELAY = 1.0     # 并发模式下每个请求的最小间隔（秒），降低风控概率

# ==========================================================================


# ----------------------------- X-Bogus（纯 Python，已验证） -----------------------------
class XBogus:
    """经典 X-Bogus 算法纯 Python 实现（来源：JohnserfSeed/TikTokDownload，Apache-2.0）。"""

    def __init__(self) -> None:
        self.character = "Dkdpgh4ZKsQB80/Mfvw36XI1R25-WUAlEi7NLboqYTOPuzmFjJnryx9HVGcaStCe="

    def md5_str_to_array(self, s):
        """将十六进制摘要串转为字节列表；非十六进制串退回 UTF-8 字节。"""
        if isinstance(s, str):
            try:
                return list(bytes.fromhex(s))
            except ValueError:
                return list(s.encode("utf-8"))
        return list(s)

    def md5(self, input_data):
        if isinstance(input_data, str):
            data = input_data.encode("utf-8")
        else:
            data = bytes(input_data)
        return hashlib.md5(data).hexdigest()

    def md5_encrypt(self, url_path):
        return self.md5_str_to_array(self.md5(self.md5_str_to_array(self.md5(url_path))))

    def encoding_conversion(self, a, b, c, e, d, t, f, r, n, o, i, _, x, u, s, l, v, h, p):
        y = [a]
        y.append(int(i))
        y.extend([b, _, c, x, e, u, d, s, t, l, f, v, r, h, n, p, o])
        return bytes(y).decode("ISO-8859-1")

    def encoding_conversion2(self, a, b, c):
        return chr(a) + chr(b) + c

    def rc4_encrypt(self, key, data):
        S = list(range(256))
        j = 0
        out = bytearray()
        for i in range(256):
            j = (j + S[i] + key[i % len(key)]) % 256
            S[i], S[j] = S[j], S[i]
        i = j = 0
        for byte in data:
            i = (i + 1) % 256
            j = (j + S[i]) % 256
            S[i], S[j] = S[j], S[i]
            out.append(byte ^ S[(S[i] + S[j]) % 256])
        return out

    def calculation(self, a1, a2, a3):
        x1 = (a1 & 255) << 16
        x2 = (a2 & 255) << 8
        x3 = x1 | x2 | a3
        return (self.character[(x3 & 16515072) >> 18] + self.character[(x3 & 258048) >> 12]
                + self.character[(x3 & 4032) >> 6] + self.character[x3 & 63])

    def getXBogus(self, url_path):
        array1 = self.md5_str_to_array("d88201c9344707acde7261b158656c0e")
        array2 = self.md5_str_to_array(self.md5(self.md5_str_to_array("d41d8cd98f00b204e9800998ecf8427e")))
        url_path_array = self.md5_encrypt(url_path)

        timer = int(time.time())
        ct = 536919696
        new_array = [
            64, 0.00390625, 1, 8,
            url_path_array[14], url_path_array[15], array2[14], array2[15], array1[14], array1[15],
            timer >> 24 & 255, timer >> 16 & 255, timer >> 8 & 255, timer & 255,
            ct >> 24 & 255, ct >> 16 & 255, ct >> 8 & 255, ct & 255
        ]
        xor_result = new_array[0]
        for i in range(1, len(new_array)):
            b = new_array[i]
            if isinstance(b, float):
                b = int(b)
            xor_result ^= b
        new_array.append(xor_result)

        idx = 0
        array3, array4 = [], []
        while idx < len(new_array):
            array3.append(new_array[idx])
            try:
                array4.append(new_array[idx + 1])
            except IndexError:
                pass
            idx += 2
        merge_array = array3 + array4

        garbled = self.encoding_conversion2(
            2, 255,
            self.rc4_encrypt("ÿ".encode("ISO-8859-1"),
                             self.encoding_conversion(*merge_array).encode("ISO-8859-1")).decode("ISO-8859-1"))

        xb_ = ""
        idx = 0
        while idx < len(garbled):
            xb_ += self.calculation(ord(garbled[idx]), ord(garbled[idx + 1]), ord(garbled[idx + 2]))
            idx += 3
        return xb_


# X-Bogus 实例无状态，复用单例避免每次请求重复实例化
_XB = XBogus()

def gen_x_bogus(query: str) -> str:
    """对请求 query 字符串生成 X-Bogus（复用单例）。"""
    return _XB.getXBogus(query)


# ----------------------------- AES 解密 -----------------------------

def _resolve_aes_key_iv():
    if AES_MODE.upper() == "CBC":
        key = base64.b64decode(AES_KEY_CBC)
        iv = base64.b64decode(AES_IV_CBC)
    elif AES_MODE.upper() == "CFB":
        key = AES_KEY_CFB.encode("utf-8")
        iv = AES_IV_CFB.encode("utf-8")
    else:
        raise ValueError("AES_MODE 仅支持 CBC 或 CFB")
    return key, iv


def aes_decrypt(cipher_b64: str) -> str:
    """Base64 解码后用 AES 解密，返回明文字符串。"""
    key, iv = _resolve_aes_key_iv()
    raw = base64.b64decode(cipher_b64)
    if AES_MODE.upper() == "CBC":
        cipher = AES.new(key, AES.MODE_CBC, iv)
        decrypted = cipher.decrypt(raw)
        try:
            decrypted = unpad(decrypted, AES.block_size)
        except Exception:
            pass  # 非标准填充时保留原样
        return decrypted.decode("utf-8", errors="ignore")
    else:  # CFB
        cipher = AES.new(key, AES.MODE_CFB, iv, segment_size=128)
        decrypted = cipher.decrypt(raw)
        return decrypted.decode("utf-8", errors="ignore")


# ----------------------------- 请求与参数拼接 -----------------------------

def parse_cookie(cookie_str: str) -> dict:
    out = {}
    for part in cookie_str.split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        k, v = part.split("=", 1)
        out[k.strip()] = v.strip()
    return out


def extract_ms_token(cookie_str: str) -> str:
    cookies = parse_cookie(cookie_str)
    return cookies.get("msToken", "")


def build_body(keyword_list, start_date, end_date) -> str:
    body = {
        "keyword_list": keyword_list,
        "start_date": start_date,
        "end_date": end_date,
        "app_name": APP_NAME,
        "region": REGION,
    }
    return json.dumps(body, separators=(",", ":"), ensure_ascii=False)


def build_request(keyword_list, start_date, end_date):
    """返回 (url, headers, cookies, data)。"""
    body_str = build_body(keyword_list, start_date, end_date)
    ms_token = extract_ms_token(COOKIE)

    query_pairs = [f"msToken={ms_token}"]
    query_without_xb = "&".join(query_pairs)

    xb = gen_x_bogus(query_without_xb)
    full_query = query_without_xb + f"&X-Bogus={xb}"

    headers = {
        "User-Agent": UA,
        "Content-Type": "application/json;charset=UTF-8",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://trendinsight.oceanengine.com/",
        "Origin": "https://trendinsight.oceanengine.com",
    }
    cookies = parse_cookie(COOKIE)
    url = ENDPOINT + "?" + full_query
    return url, headers, cookies, body_str.encode("utf-8")


def fetch_trend(keyword_list, start_date, end_date):
    """请求一批关键词，返回解密后的 JSON（dict/list）。"""
    url, headers, cookies, data = build_request(keyword_list, start_date, end_date)
    resp = requests.post(url, headers=headers, cookies=cookies, data=data, timeout=TIMEOUT)
    resp.raise_for_status()
    ctype = resp.headers.get("Content-Type", "")
    stripped = resp.text.lstrip()
    if "json" not in ctype or not stripped.startswith(("{", "[")):
        # 服务端未返回 JSON：通常是登录态失效触发的反爬/风控 HTML 页面（如 argus 挑战页）
        snippet = stripped[:200].replace("\r", " ").replace("\n", " ")
        raise RuntimeError(
            "接口返回的不是 JSON（疑似反爬/风控页面），通常是 COOKIE 无效或过期、"
            "缺少有效 msToken。请更新脚本顶部 COOKIE 为浏览器登录 trendinsight.oceanengine.com "
            "后的真实值，再重新运行。\n响应片段：" + snippet
        )
    j = resp.json()
    if j.get("status") not in (0, None):
        # 部分接口 status 字段语义不同，这里仅做提示
        print(f"  [警告] API 返回 status={j.get('status')}, msg={j.get('msg')}", file=sys.stderr)
    cipher = j.get("data")
    if not cipher:
        raise RuntimeError(f"响应缺少 data 字段，原始响应：{json.dumps(j, ensure_ascii=False)[:500]}")
    plain = aes_decrypt(cipher)
    try:
        return json.loads(plain)
    except json.JSONDecodeError:
        raise RuntimeError(f"解密后不是合法 JSON，请检查 AES_MODE/key/iv 是否正确。明文片段：{plain[:200]}")


# ----------------------------- 结果扁平化 -----------------------------

def flatten_records(records):
    """将解密后的 JSON 自适应地展开为 (keyword, date, value) 行列表。"""
    rows = []
    if isinstance(records, dict):
        items = None
        for k in ("data", "list", "items", "result", "trend_list", "hot_list"):
            if k in records and isinstance(records[k], list):
                items = records[k]
                break
        if items is None:
            items = [records]
    elif isinstance(records, list):
        items = records
    else:
        items = [{"raw": records}]

    for item in items:
        if not isinstance(item, dict):
            rows.append({"keyword": "", "date": "", "value": json.dumps(item, ensure_ascii=False)})
            continue
        keyword = item.get("keyword") or item.get("word") or item.get("name") or item.get("key") or ""
        # search_hot_list 是巨量算数常见的嵌套键
        series = None
        for k in ("search_hot_list", "index_list", "list", "data", "trend_list", "items", "point_list"):
            if k in item and isinstance(item[k], list):
                series = item[k]
                break
        if series:
            for pt in series:
                if isinstance(pt, dict):
                    date = pt.get("date") or pt.get("time") or pt.get("datetime") or pt.get("day") or ""
                    value = pt.get("value") or pt.get("index") or pt.get("count") or pt.get("val") or ""
                    rows.append({"keyword": keyword, "date": date, "value": value})
                else:
                    rows.append({"keyword": keyword, "date": "", "value": pt})
        else:
            # 无时间序列结构，整条作为原始 JSON 落表，避免丢数据
            rows.append({"keyword": keyword, "date": "", "value": json.dumps(item, ensure_ascii=False)})
    return rows


def _is_zero(v):
    try:
        return int(v) == 0
    except (ValueError, TypeError):
        return False


# ----------------------------- Excel 增量写出 -----------------------------

def write_excel(rows, filepath, canonical_keywords=None):
    """按 关键词(归一化)+日期 去重，把新行增量追加写入 xlsx。"""
    keyword_norm_map = {}
    kw_order = {}
    if canonical_keywords:
        for i, kw in enumerate(canonical_keywords):
            norm = kw.lower().replace(" ", "")
            keyword_norm_map[norm] = kw
            kw_order[norm] = i

    existing = set()
    all_rows = []
    temp_path = filepath + ".tmp.xlsx"

    source_path = None
    if os.path.exists(filepath):
        source_path = filepath
    elif os.path.exists(temp_path):
        source_path = temp_path

    if source_path:
        wb = load_workbook(source_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1]:
                raw_kw = str(row[0]); dt = str(row[1]); idx = row[2]
                norm = raw_kw.lower().replace(" ", "")
                kw = keyword_norm_map.get(norm, raw_kw)
                all_rows.append([kw, dt, idx])
                existing.add((kw, dt))
        wb.close()

    new_count = 0
    for row in rows:
        raw_kw = row["keyword"]
        norm = raw_kw.lower().replace(" ", "") if raw_kw else ""
        kw = keyword_norm_map.get(norm, raw_kw)
        dt = str(row["date"])
        key = (kw, dt)
        if key in existing:
            continue
        all_rows.append([kw, dt, row.get("value", "")])
        existing.add(key)
        new_count += 1

    def sort_key(r):
        norm = r[0].lower().replace(" ", "")
        return (kw_order.get(norm, 999), r[1])
    all_rows.sort(key=sort_key)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.append(["关键词", "日期", "关键词搜索指数"])
    for r in all_rows:
        out_ws.append(r)
    total = len(all_rows)
    try:
        out_wb.save(filepath)
        print(f"\n→ Excel 已保存: {filepath} (共 {total} 行, 本次新增 {new_count} 行)")
    except PermissionError:
        out_wb.save(temp_path)
        print(f"\n⚠ 原文件被占用，已保存到临时文件: {temp_path}")
    return new_count


def resolve_keywords_path():
    """关键词文件查找顺序：项目目录 keywords.txt → skill 根目录 keywords.txt → 无（退回内联）。"""
    proj_kw = os.path.join(KEYWORDS_DIR, "keywords.txt")
    if os.path.exists(proj_kw):
        return proj_kw
    skill_kw = os.path.join(SKILL_ROOT, "keywords.txt")
    if os.path.exists(skill_kw):
        return skill_kw
    return None


def load_keywords(path):
    keywords = []
    if not os.path.exists(path):
        return keywords
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                keywords.append(line)
    print(f"→ 从 {os.path.basename(path)} 加载 {len(keywords)} 个关键词")
    return keywords


# ----------------------------- 主流程 -----------------------------

def fetch_trend_with_retry(keyword_list, start_date, end_date, retries=MAX_RETRIES_DEFAULT):
    """带指数退避重试的批次请求。风控页面（非 JSON / 缺 data / 解密非法）不重试，需更换 Cookie。"""
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            return fetch_trend(keyword_list, start_date, end_date)
        except (requests.RequestException, RuntimeError) as ex:
            msg = str(ex)
            if "不是 JSON" in msg or "缺少 data 字段" in msg or "解密后不是合法" in msg:
                raise  # 业务/风控错误，重试无意义
            last_exc = ex
            if attempt < retries:
                wait = 2 ** attempt
                print(f"  [重试 {attempt}/{retries}] 请求失败: {ex}；{wait}s 后重试", file=sys.stderr)
                time.sleep(wait)
    raise last_exc


def load_state():
    """读取增量状态：{keyword: last_date_str}。"""
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return {}


def save_state(state):
    """原子写回增量状态（先写临时文件再 os.replace）。"""
    os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
    tmp = STATE_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    os.replace(tmp, STATE_FILE)


def per_keyword_window(kw, start, end, days, state, full):
    """
    返回该关键词本次应请求的 (start_date, end_date)；若已是最新则返回 None（跳过）。
    - full / 显式 start,end：忽略增量，走 resolve_window
    - state 有该词记录：从 last_date+1 起到今天（增量）
    - 否则：默认回推 HISTORY_DAYS 天
    """
    if full or start or end:
        return resolve_window(start, end, days)
    last = state.get(kw)
    if last:
        s = datetime.strptime(last, "%Y%m%d") + timedelta(days=1)
        e = datetime.now()
        if s.date() > e.date():
            return None  # 已是最新
        return s.strftime("%Y%m%d"), e.strftime("%Y%m%d")
    return resolve_window(start, end, days)


def chunked(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def resolve_window(start, end, days):
    if start and end:
        return start, end
    n_days = days if days else HISTORY_DAYS
    e = datetime.now()
    s = e - timedelta(days=n_days)
    return s.strftime("%Y%m%d"), e.strftime("%Y%m%d")


def main(keywords=None, start=None, end=None, days=None, concurrency=CONCURRENCY_DEFAULT,
         full=False, retries=MAX_RETRIES_DEFAULT):
    global COOKIE
    if COOKIE == "PASTE_YOUR_COOKIE_HERE":
        sys.exit("请先在脚本顶部 COOKIE 配置块填入浏览器登录后的真实 Cookie，再运行。")

    kws = keywords
    if not kws:
        kw_path = resolve_keywords_path()
        if kw_path:
            kws = load_keywords(kw_path)
    if not kws:
        kws = KEYWORDS
    if not kws:
        sys.exit("未配置关键词：请在项目目录或 skill 根目录放 keywords.txt（每行一个），"
                 "或在脚本顶部 KEYWORDS 填写，或用 --keywords 指定。")

    excel_file = os.path.join(OUTPUT_DIR, "keyword_index.xlsx")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    state = load_state()  # 始终加载；full 仅影响窗口计算，不丢失其他词的记录

    # 计算每个关键词本次的增量窗口
    jobs = []  # (kw, start_date, end_date)
    skipped = 0
    for kw in kws:
        w = per_keyword_window(kw, start, end, days, state, full)
        if w is None:
            print(f"  ⊘ {kw}: 已是最新(last={state.get(kw)})，跳过")
            skipped += 1
            continue
        jobs.append((kw, w[0], w[1]))

    if not jobs:
        print("\n✓ 所有关键词均已是最新（用 --full 可强制全量重采）")
        return

    print("=" * 60)
    print("  巨量算数关键词指数采集（直连 API + AES 解密）")
    print(f"  时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  关键词数: {len(kws)} | 本次待采集: {len(jobs)} | 跳过(已最新): {skipped}")
    print(f"  并发: {concurrency} | 重试: {retries} | 全量: {full}")
    print("=" * 60)

    def fetch_one(kw, s, e):
        try:
            records = fetch_trend_with_retry([kw], s, e, retries)
            rows = flatten_records(records)
            if not KEEP_ZERO_RECORDS:
                rows = [r for r in rows if not _is_zero(r["value"])]
            dates = [r["date"] for r in rows if r.get("date")]
            last = max(dates) if dates else None
            return kw, rows, last, None
        except Exception as ex:
            return kw, None, None, str(ex)

    all_rows = []
    new_last = {}  # kw -> last_date，待写回 state

    if concurrency <= 1:
        for i, (kw, s, e) in enumerate(jobs, 1):
            print(f"\n[{i}/{len(jobs)}] {kw}  窗口 {s}~{e}")
            _, rows, last, err = fetch_one(kw, s, e)
            if err:
                print(f"  [错误] {kw} 失败：{err}", file=sys.stderr)
            else:
                all_rows.extend(rows)
                if last:
                    new_last[kw] = last
                print(f"  -> 得到 {len(rows)} 行" + (f"，最新 {last}" if last else ""))
            time.sleep(DELAY)
    else:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=concurrency) as ex:
            futs = {}
            for kw, s, e in jobs:
                time.sleep(PER_REQUEST_DELAY)  # 控制提交频率，降低风控概率
                futs[ex.submit(fetch_one, kw, s, e)] = (kw, s, e)
            done = 0
            for fut in as_completed(futs):
                kw, s, e = futs[fut]
                done += 1
                _, rows, last, err = fut.result()
                if err:
                    print(f"  [错误] {kw} 失败：{err}", file=sys.stderr)
                else:
                    all_rows.extend(rows)
                    if last:
                        new_last[kw] = last
                    print(f"[{done}/{len(jobs)}] {kw} -> {len(rows)} 行" + (f"，最新 {last}" if last else ""))

    if not all_rows:
        print("\n❌ 未获取到任何数据（可能 Cookie 过期 / 接口变更 / 被风控）")
        sys.exit(1)

    write_excel(all_rows, excel_file, canonical_keywords=kws)

    # 写回增量状态
    if new_last:
        state.update(new_last)
        save_state(state)
        print(f"→ 已更新增量状态：{len(new_last)} 个关键词的末日记录")

    print(f"\n✅ 完成 — Excel: {excel_file}")


def verify(keywords=None, start=None, end=None, days=None):
    global COOKIE
    if COOKIE == "PASTE_YOUR_COOKIE_HERE":
        sys.exit("请先在脚本顶部 COOKIE 配置块填入浏览器登录后的真实 Cookie，再运行。")

    kws = (keywords if keywords else KEYWORDS)[:BATCH_SIZE]
    if not kws:
        sys.exit("未配置关键词。")
    start_date, end_date = resolve_window(start, end, days)
    print("[VERIFY] 使用关键词：", kws)
    print(f"[VERIFY] msToken 是否为空：{extract_ms_token(COOKIE) == ''}")
    print(f"[VERIFY] X-Bogus 样例：{gen_x_bogus('msToken=' + extract_ms_token(COOKIE))}")
    try:
        records = fetch_trend(kws, start_date, end_date)
        rows = flatten_records(records)
        print(f"[VERIFY] 解密成功，展开 {len(rows)} 行。样例：")
        for r in rows[:5]:
            print("   ", r)
        print("[VERIFY] 完整解密 JSON（前 800 字）：")
        print(json.dumps(records, ensure_ascii=False, indent=2)[:800])
    except Exception as e:
        print(f"[VERIFY] 失败：{e}", file=sys.stderr)


def parse_args():
    p = argparse.ArgumentParser(description="巨量算数关键词搜索指数采集（直连 API + AES 解密）")
    p.add_argument("--verify", action="store_true", help="试跑单个批次并打印解密结果")
    p.add_argument("--keywords", nargs="+", help="覆盖配置区的关键词列表")
    p.add_argument("--start", help="开始日期 YYYYMMDD（指定后忽略增量，走显式窗口）")
    p.add_argument("--end", help="结束日期 YYYYMMDD")
    p.add_argument("--days", type=int, help="相对今天回推的天数（覆盖 HISTORY_DAYS）")
    p.add_argument("--concurrency", type=int, default=CONCURRENCY_DEFAULT,
                   help=f"并发请求数（默认 {CONCURRENCY_DEFAULT}，串行最稳；>1 可加速首跑但注意风控）")
    p.add_argument("--full", action="store_true",
                   help="强制全量采集，忽略增量状态（默认只在已有数据后补新日期）")
    p.add_argument("--retries", type=int, default=MAX_RETRIES_DEFAULT,
                   help=f"网络 / 5xx 瞬态错误重试次数（默认 {MAX_RETRIES_DEFAULT}；风控页面不重试）")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    if args.verify:
        verify(args.keywords, args.start, args.end, args.days)
    else:
        main(args.keywords, args.start, args.end, args.days,
             concurrency=args.concurrency, full=args.full, retries=args.retries)
